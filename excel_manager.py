import os
import logging
from typing import Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelBackupManager:
    """
    Класс для управления резервным копированием данных пользователей в Excel файл.

    Attributes:
        file_path (str): Путь к Excel файлу
        workbook (Workbook): Объект рабочей книги Excel
        worksheet: Активный лист для работы с данными
        headers (list): Заголовки столбцов
    """

    def __init__(self, file_path: str = "backup.xlsx"):
        """
        Инициализация менеджера Excel резервного копирования.

        Args:
            file_path (str): Путь к файлу резервной копии (по умолчанию "backup.xlsx")
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.headers = ["ID", "Full Name", "Summ", "Card Number", "Birthday"]

        # Создаем директорию если она не существует
        self._ensure_directory_exists()

        # Инициализируем файл
        self._initialize_workbook()

    def _ensure_directory_exists(self) -> None:
        """Создает директорию для файла резервной копии если она не существует."""
        directory = os.path.dirname(self.file_path)
        if directory and not os.path.exists(directory):
            try:
                os.makedirs(directory, exist_ok=True)
                logger.info(f"Создана директория: {directory}")
            except OSError as e:
                logger.error(f"Ошибка создания директории {directory}: {e}")
                raise

    def _initialize_workbook(self) -> None:
        """Инициализирует рабочую книгу Excel."""
        try:
            if os.path.exists(self.file_path):
                # Загружаем существующий файл
                self.workbook = load_workbook(self.file_path)
                if "users" not in self.workbook.sheetnames:
                    # Создаем лист users если его нет
                    self.worksheet = self.workbook.create_sheet("users")
                    self._write_headers()
                else:
                    self.worksheet = self.workbook["users"]
                logger.info(f"Загружен существующий файл: {self.file_path}")
            else:
                # Создаем новый файл
                self.workbook = Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = "users"
                self._write_headers()
                logger.info(f"Создан новый файл резервной копии: {self.file_path}")
        except Exception as e:
            logger.error(f"Ошибка инициализации Excel файла: {e}")
            raise

    def _write_headers(self) -> None:
        """Записывает заголовки столбцов в Excel файл."""
        try:
            for col_num, header in enumerate(self.headers, 1):
                cell = self.worksheet.cell(row=1, column=col_num, value=header)

                # Стилизация заголовков
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

                # Автоматическая ширина столбцов
                column_letter = get_column_letter(col_num)
                self.worksheet.column_dimensions[column_letter].width = 15

            logger.info("Заголовки успешно записаны в Excel файл")
        except Exception as e:
            logger.error(f"Ошибка записи заголовков: {e}")
            raise

    def ensure_file_exists(self) -> bool:
        """
        Проверяет существование файла и создает его с заголовками если необходимо.

        Returns:
            bool: True если файл существует или успешно создан
        """
        try:
            if not os.path.exists(self.file_path):
                self._initialize_workbook()
                self.save()
                logger.info(f"Файл резервной копии создан: {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"Ошибка создания файла резервной копии: {e}")
            return False

    def add_user(self, user_data: Dict) -> bool:
        """
        Добавляет запись о пользователе в Excel файл.

        Args:
            user_data (dict): Данные пользователя в формате:
                {
                    "id": int,
                    "full_name": str,
                    "summ": float,
                    "card_number": int,
                    "birthday": str  # формат YYYY-MM-DD
                }

        Returns:
            bool: True если запись успешно добавлена
        """
        try:
            # Проверяем корректность данных
            required_fields = ["id", "full_name", "summ", "card_number", "birthday"]
            for field in required_fields:
                if field not in user_data:
                    raise ValueError(f"Отсутствует обязательное поле: {field}")

            # Находим следующую пустую строку
            next_row = self.worksheet.max_row + 1

            # Записываем данные в соответствующие столбцы
            self.worksheet.cell(row=next_row, column=1, value=user_data["id"])
            self.worksheet.cell(row=next_row, column=2, value=user_data["full_name"])
            self.worksheet.cell(row=next_row, column=3, value=user_data["summ"])
            self.worksheet.cell(row=next_row, column=4, value=user_data["card_number"])

            # Обрабатываем дату рождения
            birthday_cell = self.worksheet.cell(row=next_row, column=5, value=user_data["birthday"])
            birthday_cell.number_format = 'YYYY-MM-DD'

            # Сохраняем изменения
            self.save()

            logger.info(f"Пользователь {user_data['full_name']} (ID: {user_data['id']}) добавлен в резервную копию")
            return True

        except Exception as e:
            logger.error(f"Ошибка добавления пользователя в Excel: {e}")
            return False

    def save(self) -> bool:
        """
        Сохраняет изменения в Excel файл.

        Returns:
            bool: True если файл успешно сохранен
        """
        try:
            if self.workbook:
                self.workbook.save(self.file_path)
                logger.debug("Excel файл сохранен")
                return True
            return False
        except Exception as e:
            logger.error(f"Ошибка сохранения Excel файла: {e}")
            return False

    def close(self) -> None:
        """Корректно закрывает Excel файл."""
        try:
            if self.workbook:
                self.save()
                self.workbook.close()
                logger.info("Excel файл закрыт")
        except Exception as e:
            logger.error(f"Ошибка при закрытии Excel файла: {e}")

    def get_existing_ids(self) -> set:
        """
        Получает множество ID пользователей, которые уже существуют в Excel файле.

        Returns:
            set: Множество существующих ID
        """
        try:
            existing_ids = set()
            if not self.worksheet or self.worksheet.max_row < 2:  # только заголовки или пустой файл
                return existing_ids

            # Проходим по столбцу ID (столбец 1) начиная со второй строки (после заголовков)
            for row in range(2, self.worksheet.max_row + 1):
                cell_value = self.worksheet.cell(row=row, column=1).value
                if cell_value is not None:
                    try:
                        existing_ids.add(int(cell_value))
                    except (ValueError, TypeError):
                        # Пропускаем некорректные значения
                        continue

            logger.debug(f"Найдено {len(existing_ids)} существующих ID в Excel файле")
            return existing_ids

        except Exception as e:
            logger.error(f"Ошибка получения существующих ID из Excel: {e}")
            return set()

    def clear_backup(self) -> bool:
        """
        Очищает Excel файл, оставляя только заголовки.

        Returns:
            bool: True если файл успешно очищен
        """
        try:
            if not self.worksheet:
                logger.warning("Рабочий лист не инициализирован")
                return False

            # Оставляем только первую строку (заголовки)
            self.worksheet.delete_rows(2, self.worksheet.max_row)

            # Сохраняем изменения
            self.save()

            logger.info("Excel файл резервной копии очищен (оставлены только заголовки)")
            return True

        except Exception as e:
            logger.error(f"Ошибка очистки Excel файла: {e}")
            return False

    def get_backup_info(self) -> Dict:
        """
        Получает информацию о резервной копии.

        Returns:
            dict: Информация о файле резервной копии
        """
        try:
            return {
                "file_path": self.file_path,
                "file_exists": os.path.exists(self.file_path),
                "file_size": os.path.getsize(self.file_path) if os.path.exists(self.file_path) else 0,
                "total_records": self.worksheet.max_row - 1 if self.worksheet else 0,  # минус заголовок
                "last_modified": os.path.getmtime(self.file_path) if os.path.exists(self.file_path) else None
            }
        except Exception as e:
            logger.error(f"Ошибка получения информации о резервной копии: {e}")
            return {
                "file_path": self.file_path,
                "error": str(e)
            }


# Singleton паттерн для глобального доступа к менеджеру
_excel_manager_instance = None


def get_excel_manager(file_path: str = "backup.xlsx") -> ExcelBackupManager:
    """
    Получает экземпляр ExcelBackupManager (singleton).

    Args:
        file_path (str): Путь к файлу резервной копии

    Returns:
        ExcelBackupManager: Экземпляр менеджера
    """
    global _excel_manager_instance
    if _excel_manager_instance is None:
        _excel_manager_instance = ExcelBackupManager(file_path)
    return _excel_manager_instance
